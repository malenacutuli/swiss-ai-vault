"""
XLSX spreadsheet generator using openpyxl.

Generates Excel spreadsheets with:
- Multiple sheets
- Formatted headers
- Data tables
- Cell styling
- Formulas
"""

import os
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .base import DocumentGenerator, DocumentContent


class XLSXGenerator(DocumentGenerator):
    """Generator for Microsoft Excel (.xlsx) spreadsheets."""

    def __init__(self, content: DocumentContent):
        """
        Initialize XLSX generator.

        Args:
            content: DocumentContent with title and sections
        """
        super().__init__(content)
        self.workbook = Workbook()
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]

    def get_format_name(self) -> str:
        """Return format name."""
        return "xlsx"

    def validate_content(self) -> bool:
        """
        Validate content structure for XLSX generation.

        Returns:
            bool: True if valid

        Raises:
            ValueError: If content is invalid
        """
        if not self.content.title:
            raise ValueError("Spreadsheet title is required")

        if not self.content.sections:
            raise ValueError("At least one section is required")

        for idx, section in enumerate(self.content.sections):
            if "heading" not in section:
                raise ValueError(f"Section {idx} missing 'heading' field")

        return True

    def generate(self, output_path: str) -> str:
        """
        Generate XLSX spreadsheet.

        Args:
            output_path: Path to save the spreadsheet

        Returns:
            str: Path to generated spreadsheet
        """
        self.validate_content()

        # Create overview sheet
        self._add_overview_sheet()

        # Add data sheets for each section
        for idx, section in enumerate(self.content.sections):
            self._add_section_sheet(section, idx)

        # Save workbook
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        self.workbook.save(output_path)

        # Update metadata
        file_size = os.path.getsize(output_path)
        page_count = len(self.workbook.sheetnames)
        self.update_metadata(file_size=file_size, page_count=page_count)

        return output_path

    def _add_overview_sheet(self):
        """Add overview sheet with document info."""
        sheet = self.workbook.create_sheet("Overview", 0)

        # Title
        sheet['A1'] = self.content.title
        sheet['A1'].font = Font(size=18, bold=True, color="0070C0")
        sheet.row_dimensions[1].height = 30

        # Metadata
        if self.content.metadata:
            sheet['A3'] = "Document Information"
            sheet['A3'].font = Font(size=12, bold=True)

            row = 4
            for key, value in self.content.metadata.items():
                sheet[f'A{row}'] = str(key).replace('_', ' ').title()
                sheet[f'B{row}'] = str(value)
                sheet[f'A{row}'].font = Font(bold=True)
                row += 1

        # Section summary
        summary_row = 8
        sheet[f'A{summary_row}'] = "Sections"
        sheet[f'A{summary_row}'].font = Font(size=12, bold=True)

        # Headers
        summary_row += 1
        sheet[f'A{summary_row}'] = "Section"
        sheet[f'B{summary_row}'] = "Type"
        sheet[f'C{summary_row}'] = "Sheet Name"
        for cell in [sheet[f'A{summary_row}'], sheet[f'B{summary_row}'], sheet[f'C{summary_row}']]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Section list
        for idx, section in enumerate(self.content.sections):
            summary_row += 1
            sheet[f'A{summary_row}'] = section.get("heading", f"Section {idx + 1}")
            sheet[f'B{summary_row}'] = section.get("type", "text")
            sheet[f'C{summary_row}'] = f"Sheet_{idx + 1}"

        # Auto-adjust column widths
        for column in ['A', 'B', 'C']:
            sheet.column_dimensions[column].width = 20

    def _add_section_sheet(self, section: Dict[str, Any], idx: int):
        """
        Add a sheet for a section.

        Args:
            section: Section data
            idx: Section index
        """
        sheet_name = section.get("heading", f"Section_{idx + 1}")[:31]  # Excel limit
        sheet = self.workbook.create_sheet(sheet_name)

        # Section heading
        sheet['A1'] = section.get("heading", f"Section {idx + 1}")
        sheet['A1'].font = Font(size=14, bold=True, color="0070C0")
        sheet.row_dimensions[1].height = 25

        content_type = section.get("type", "text")

        if content_type == "table":
            self._add_table_to_sheet(sheet, section)
        elif content_type in ["bullet_list", "numbered_list"]:
            self._add_list_to_sheet(sheet, section)
        else:
            self._add_text_to_sheet(sheet, section)

    def _add_table_to_sheet(self, sheet, section: Dict[str, Any]):
        """
        Add table data to a sheet.

        Args:
            sheet: Excel worksheet
            section: Section with table data
        """
        headers = section.get("headers", [])
        rows = section.get("rows", [])

        if not headers or not rows:
            return

        # Add headers
        header_row = 3
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col_idx)
            cell.value = str(header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Add data rows
        for row_idx, row_data in enumerate(rows, start=header_row + 1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = cell_value
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 15

        # Add totals if numeric data
        if section.get("add_totals", False):
            total_row = header_row + len(rows) + 1
            sheet.cell(row=total_row, column=1).value = "TOTAL"
            sheet.cell(row=total_row, column=1).font = Font(bold=True)

            for col_idx in range(2, len(headers) + 1):
                cell = sheet.cell(row=total_row, column=col_idx)
                # Try to add SUM formula
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}{header_row + 1}:{col_letter}{header_row + len(rows)})"
                cell.font = Font(bold=True)

    def _add_list_to_sheet(self, sheet, section: Dict[str, Any]):
        """
        Add list items to a sheet.

        Args:
            sheet: Excel worksheet
            section: Section with list items
        """
        items = section.get("items", [])

        # Add items starting at row 3
        for idx, item in enumerate(items, start=3):
            sheet[f'A{idx}'] = str(item)
            if section.get("type") == "numbered_list":
                sheet[f'A{idx}'] = f"{idx - 2}. {item}"

        # Auto-adjust column width
        sheet.column_dimensions['A'].width = 50

    def _add_text_to_sheet(self, sheet, section: Dict[str, Any]):
        """
        Add text content to a sheet.

        Args:
            sheet: Excel worksheet
            section: Section with text content
        """
        content = section.get("content", "")

        if isinstance(content, list):
            for idx, item in enumerate(content, start=3):
                sheet[f'A{idx}'] = str(item)
        else:
            sheet['A3'] = str(content)

        # Auto-adjust column width
        sheet.column_dimensions['A'].width = 60
